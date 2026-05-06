"""Generate supplier-facing email drafts for failed_image orders."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
import io
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.repositories import customer_repository
from app.services.failed_reset_service import DEFAULT_SUPPLIER_EMAIL, RUBEN_EMAIL, SUPPLIER_ROUTING
from app.utils.domain_normalization import normalize_domain

logger = logging.getLogger(__name__)

MAX_DOMAINS_PER_DRAFT = 50
PREVIEW_DOMAINS_PER_DRAFT = 50


@dataclass(frozen=True)
class FailedImageDraft:
    recipient_name: str
    recipient_email: str
    subject: str
    body_html: str
    body_json: dict[str, Any]
    order_count: int
    domain_count: int
    all_domains: list[dict[str, Any]]


def get_failed_image_orders(source_session: Session, db_name: str) -> list[dict[str, Any]]:
    """Fetch current failed_image orders from one source database."""
    result = source_session.execute(text("""
        SELECT
            oo.id AS order_id,
            oo.domainId,
            d.wp_domain,
            oo.status,
            oo.addedOn,
            oo.deliveryDate,
            oo.addedBy,
            oo.customerId,
            oo.anchor1,
            oo.anchor2,
            oo.anchor3,
            oo.link1,
            oo.link2,
            oo.link3,
            GROUP_CONCAT(DISTINCT dl.labelId ORDER BY dl.labelId SEPARATOR ', ') AS label_ids,
            GROUP_CONCAT(DISTINCT l.name ORDER BY l.name SEPARATOR ', ') AS label_names
        FROM openorder oo
        LEFT JOIN domains d ON oo.domainId = d.id
        LEFT JOIN domlabels dl ON dl.domId = d.id
        LEFT JOIN labels l ON l.id = dl.labelId
        WHERE oo.status LIKE '%failed_image%'
        GROUP BY
            oo.id, oo.domainId, d.wp_domain, oo.status, oo.addedOn, oo.deliveryDate,
            oo.addedBy, oo.customerId, oo.anchor1, oo.anchor2, oo.anchor3,
            oo.link1, oo.link2, oo.link3
        ORDER BY oo.addedBy, d.wp_domain, oo.id
    """))

    rows: list[dict[str, Any]] = []
    for row in result.fetchall():
        rows.append({
            "db": db_name,
            "order_id": row[0],
            "domainId": row[1],
            "wp_domain": row[2] or "",
            "status": row[3] or "",
            "added_on": str(row[4] or ""),
            "delivery_date": str(row[5] or ""),
            "added_by": row[6] or 0,
            "customer_id": row[7] or 0,
            "anchor1": row[8] or "",
            "anchor2": row[9] or "",
            "anchor3": row[10] or "",
            "link1": row[11] or "",
            "link2": row[12] or "",
            "link3": row[13] or "",
            "label_ids": row[14] or "",
            "label_names": row[15] or "",
        })
    return rows


def enrich_failed_image_orders(source_session: Session, orders: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Add customer data to failed_image rows."""
    customer_ids = sorted({o["customer_id"] for o in orders if o.get("customer_id")})
    customers = customer_repository.get_customers_by_ids(source_session, customer_ids)

    enriched: list[dict[str, Any]] = []
    for order in orders:
        customer = customers.get(order.get("customer_id"))
        enriched.append({
            **order,
            "customer_name": customer.name if customer else "",
        })
    return enriched


def build_failed_image_owner_drafts(orders: list[dict[str, Any]]) -> list[FailedImageDraft]:
    """Build one draft per supplier: JG labels to Hugo, all others to Stefan."""
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for order in orders:
        grouped[_resolve_supplier_email(order.get("label_names", ""))].append(order)

    drafts: list[FailedImageDraft] = []
    for supplier_email, supplier_orders in grouped.items():
        recipient_name = "Hugo" if "hugo" in supplier_email.lower() else "Stefan"
        supplier_domains = summarize_orders_by_domain(supplier_orders)
        preview_domains = supplier_domains[:PREVIEW_DOMAINS_PER_DRAFT]
        total_order_count = sum(int(d.get("order_count") or 0) for d in supplier_domains)
        subject = f"Failed image: {len(supplier_domains)} domein(en)"
        body_html = build_failed_image_owner_html(
            recipient_name,
            preview_domains,
            total_domains=len(supplier_domains),
            total_orders=total_order_count,
        )
        drafts.append(FailedImageDraft(
            recipient_name=recipient_name,
            recipient_email=supplier_email,
            subject=subject,
            body_html=body_html,
            body_json={
                "type": "failed_image_supplier_draft",
                "cc": RUBEN_EMAIL,
                "reply_to": RUBEN_EMAIL,
                "preview_domain_count": len(preview_domains),
                "total_domains": len(supplier_domains),
                "total_orders": total_order_count,
                "domains": [
                    {
                        "db": o.get("db", ""),
                        "wp_domain": o.get("wp_domain", ""),
                        "customer_name": o.get("customer_name", ""),
                        "label_names": o.get("label_names", ""),
                        "order_count": o.get("order_count", 0),
                        "order_ids": o.get("order_ids", []),
                    }
                    for o in preview_domains
                ],
            },
            order_count=total_order_count,
            domain_count=len(supplier_domains),
            all_domains=supplier_domains,
        ))

    return sorted(drafts, key=lambda draft: draft.recipient_email.lower())


def _resolve_supplier_email(label_names: str) -> str:
    for label_key, email in SUPPLIER_ROUTING.items():
        if label_key.lower() in (label_names or "").lower():
            return email
    return DEFAULT_SUPPLIER_EMAIL


def summarize_orders_by_domain(orders: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for order in orders:
        key = normalize_domain(order.get("wp_domain", "")) or str(order.get("domainId") or order.get("order_id"))
        domain = grouped.setdefault(
            key,
            {
                **order,
                "order_ids": [],
                "customers": set(),
                "anchors": set(),
                "links": set(),
                "databases": set(),
                "labels": set(),
            },
        )
        domain["order_ids"].append(order.get("order_id"))
        if order.get("customer_name"):
            domain["customers"].add(order["customer_name"])
        if order.get("anchor1"):
            domain["anchors"].add(order["anchor1"])
        if order.get("link1"):
            domain["links"].add(order["link1"])
        if order.get("db"):
            domain["databases"].add(order["db"])
        if order.get("label_names"):
            domain["labels"].add(order["label_names"])

    result = []
    for domain in grouped.values():
        domain["order_count"] = len(domain["order_ids"])
        domain["customer_name"] = ", ".join(sorted(domain["customers"]))
        domain["anchor1"] = ", ".join(sorted(domain["anchors"])[:3])
        domain["link1"] = ", ".join(sorted(domain["links"])[:3])
        domain["db"] = ", ".join(sorted(domain["databases"]))
        domain["label_names"] = ", ".join(sorted(domain["labels"]))
        result.append(domain)
    return sorted(result, key=lambda item: item.get("wp_domain", ""))


def build_failed_image_owner_html(
    recipient_name: str,
    domains: list[dict[str, Any]],
    total_domains: int | None = None,
    total_orders: int | None = None,
) -> str:
    rows_html = ""
    for domain in domains:
        rows_html += (
            "<tr>"
            f"<td>{domain.get('wp_domain', '')}</td>"
            "<td>Afbeelding uploaden geblokkeerd</td>"
            f"<td>{domain.get('order_count', 0)}</td>"
            "</tr>\n"
        )

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
body {{ font-family: Arial, sans-serif; font-size: 14px; color: #333; line-height: 1.6; }}
table {{ border-collapse: collapse; width: 100%; margin: 16px 0; }}
th, td {{ border: 1px solid #ddd; padding: 8px 10px; text-align: left; font-size: 13px; }}
th {{ background: #f5f5f5; font-weight: 600; }}
</style></head><body>
<p>Hi {recipient_name},</p>

<p>Onderstaande linkbuilding-orders staan momenteel op <code>failed_image</code>.
Wil je controleren waarom de afbeelding niet geplaatst kan worden?</p>

<p>Hieronder staat een preview van de eerste {len(domains)} domein(en). De volledige lijst met {total_domains or len(domains)} domein(en) en {total_orders or sum(int(d.get('order_count') or 0) for d in domains)} order(s) zit als Excel-bijlage bij deze draft.</p>

<table>
<thead><tr><th>Website</th><th>Foutmelding</th><th>Aantal orders</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>

<p>Alvast bedankt!</p>

<p>Met vriendelijke groet,<br>
Ruben van Melsen<br>
Blauwe Monsters</p>
</body></html>"""


def save_failed_image_owner_drafts(
    audit_session: Session,
    drafts: list[FailedImageDraft],
    run_id: int | None = None,
    output_dir: Path | None = None,
) -> list[int]:
    from app.clients.audit_db import AuditEmailDraft

    draft_ids: list[int] = []
    for draft in drafts:
        xlsx_path = ""
        if output_dir:
            xlsx_dir = output_dir / "reports"
            xlsx_dir.mkdir(parents=True, exist_ok=True)
            safe_email = draft.recipient_email.replace("@", "_at_").replace(".", "_")
            xlsx_file = xlsx_dir / f"failed_image_{safe_email}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            xlsx_file.write_bytes(build_failed_image_xlsx(draft.all_domains))
            xlsx_path = str(xlsx_file)

        body_json = {
            **draft.body_json,
            "xlsx_path": xlsx_path,
        }
        row = AuditEmailDraft(
            run_id=run_id or 0,
            added_by=0,
            addedby_name=draft.recipient_name,
            addedby_email=draft.recipient_email,
            subject=draft.subject,
            body_html=draft.body_html,
            body_json=body_json,
            order_count=draft.order_count,
            created_at=datetime.utcnow(),
            send_status="draft",
        )
        audit_session.add(row)
        audit_session.flush()
        draft_ids.append(row.id)

    audit_session.commit()
    return draft_ids


def build_failed_image_xlsx(domains: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Failed image domeinen"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    headers = ["Website", "Foutmelding", "Aantal orders", "Klant(en)", "Database", "Labels", "Order IDs"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, domain in enumerate(domains, 2):
        ws.cell(row=row_idx, column=1, value=domain.get("wp_domain", ""))
        ws.cell(row=row_idx, column=2, value="Afbeelding uploaden geblokkeerd")
        ws.cell(row=row_idx, column=3, value=domain.get("order_count", 0))
        ws.cell(row=row_idx, column=4, value=domain.get("customer_name", ""))
        ws.cell(row=row_idx, column=5, value=domain.get("db", ""))
        ws.cell(row=row_idx, column=6, value=domain.get("label_names", ""))
        ws.cell(row=row_idx, column=7, value=", ".join(str(i) for i in domain.get("order_ids", [])))

    widths = [42, 34, 14, 28, 28, 28, 32]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
