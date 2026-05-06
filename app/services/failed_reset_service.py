"""Reset failed orders to pending, categorize by failure type, and generate supplier drafts.

Failure categories:
- failed_replace:  Always reset to pending. After 5 consecutive fails per order,
                   send notification to ruben.van.melsen@blauwemonsters.nl.
- failed_domain / failed_domain_category / failed_image:
                   Reset to pending on first occurrence. On repeat failure,
                   generate a supplier email draft (NOT sent automatically)
                   grouped by label: JG Webmarketing → hugo@jgwebmarketing.nl,
                   others → stefan@hostingindustries.nl.
                   CC + Reply-To: ruben.van.melsen@blauwemonsters.nl.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import text
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

RUBEN_EMAIL = "ruben.van.melsen@blauwemonsters.nl"

SUPPLIER_ROUTING: dict[str, str] = {
    "JG Webmarketing": "hugo@jgwebmarketing.nl",
    "JG webmarketing": "hugo@jgwebmarketing.nl",
}
DEFAULT_SUPPLIER_EMAIL = "stefan@hostingindustries.nl"

REPLACE_THRESHOLD = 5


def _is_replace_failure(status: str) -> bool:
    return "failed_replace" in (status or "")


def _is_connection_or_image_failure(status: str) -> bool:
    s = status or ""
    return any(t in s for t in ("failed_domain", "failed_image"))


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def get_failed_orders_enriched(session: Session) -> list[dict]:
    result = session.execute(text("""
        SELECT
            oo.id AS order_id,
            oo.domainId,
            d.wp_domain,
            oo.status,
            oo.addedOn,
            oo.addedBy,
            GROUP_CONCAT(DISTINCT dl.labelId ORDER BY dl.labelId SEPARATOR ', ') AS label_ids,
            GROUP_CONCAT(DISTINCT l.name ORDER BY l.name SEPARATOR ', ') AS label_names
        FROM openorder oo
        LEFT JOIN domains d ON oo.domainId = d.id
        LEFT JOIN domlabels dl ON dl.domId = d.id
        LEFT JOIN labels l ON l.id = dl.labelId
        WHERE oo.status LIKE '%failed%'
        GROUP BY oo.id, oo.domainId, d.wp_domain, oo.status, oo.addedOn, oo.addedBy
    """))
    rows = []
    for r in result.fetchall():
        rows.append({
            "order_id": r[0],
            "domainId": r[1],
            "wp_domain": r[2] or "",
            "status": r[3] or "",
            "addedOn": str(r[4] or ""),
            "addedBy": r[5],
            "label_ids": r[6] or "",
            "label_names": r[7] or "",
        })
    return rows


def get_failed_orders_simple(session: Session) -> list[dict]:
    result = session.execute(text("""
        SELECT
            oo.id AS order_id,
            oo.domainId,
            d.wp_domain,
            oo.status,
            oo.addedOn,
            oo.addedBy
        FROM openorder oo
        LEFT JOIN domains d ON oo.domainId = d.id
        WHERE oo.status LIKE '%failed%'
    """))
    rows = []
    for r in result.fetchall():
        rows.append({
            "order_id": r[0],
            "domainId": r[1],
            "wp_domain": r[2] or "",
            "status": r[3] or "",
            "addedOn": str(r[4] or ""),
            "addedBy": r[5],
            "label_ids": "",
            "label_names": "",
        })
    return rows


def get_failed_orders(session: Session) -> list[dict]:
    try:
        return get_failed_orders_enriched(session)
    except Exception:
        logger.warning("Enriched query failed (labels table may not exist), using simple query")
        session.rollback()
        return get_failed_orders_simple(session)


# ---------------------------------------------------------------------------
# Audit history helpers
# ---------------------------------------------------------------------------

def _get_reset_count_for_order(audit_session: Session, db_name: str, order_id: int) -> int:
    """Count how many times an order has been reset (non-dry-run) in the audit log."""
    from app.clients.audit_db import AuditLog
    from sqlalchemy import and_

    count = (
        audit_session.query(AuditLog)
        .filter(
            and_(
                AuditLog.level == "INFO",
                AuditLog.message.like(f"[{db_name}] Reset order {order_id} %"),
            )
        )
        .count()
    )
    return count


def _get_previously_reset_ids(audit_session: Session, db_name: str, order_ids: list[int]) -> set[str]:
    from app.clients.audit_db import AuditLog
    from sqlalchemy import and_

    if not order_ids:
        return set()

    logs = (
        audit_session.query(AuditLog)
        .filter(
            and_(
                AuditLog.level == "INFO",
                AuditLog.message.like(f"[{db_name}] Reset order %"),
            )
        )
        .all()
    )

    already_reset: set[str] = set()
    for log in logs:
        ctx = log.context_json or {}
        if ctx.get("db") == db_name and ctx.get("order_id") and not ctx.get("dry_run"):
            already_reset.add(f"{db_name}:{ctx['order_id']}")

    return already_reset


# ---------------------------------------------------------------------------
# Main entry: categorize and reset
# ---------------------------------------------------------------------------

def _get_deleted_label_domain_ids(source_session: Session, expired_label_id: int = 79) -> set[int]:
    """Get all domain IDs that have the 'Deleted' label (79)."""
    result = source_session.execute(text(
        f"SELECT DISTINCT domId FROM domlabels WHERE labelId = {expired_label_id}"
    ))
    return {r[0] for r in result.fetchall()}


def reset_failed_to_pending(
    source_session: Session,
    audit_session: Session,
    db_name: str,
    run_id: int | None = None,
    dry_run: bool = False,
) -> dict:
    from app.clients.audit_db import AuditLog

    failed_orders = get_failed_orders(source_session)
    if not failed_orders:
        logger.info("[%s] No failed orders found", db_name)
        return {
            "db": db_name, "total_failed": 0, "reset": 0, "deleted": 0,
            "supplier_draft_orders": [], "replace_notify_orders": [],
        }

    order_ids = [o["order_id"] for o in failed_orders]
    previously_reset = _get_previously_reset_ids(audit_session, db_name, order_ids)
    deleted_domain_ids = _get_deleted_label_domain_ids(source_session)

    to_reset: list[dict] = []
    to_delete: list[dict] = []
    replace_always_reset: list[dict] = []
    replace_notify: list[dict] = []
    supplier_draft_orders: list[dict] = []

    for o in failed_orders:
        key = f"{db_name}:{o['order_id']}"
        tagged = {**o, "db": db_name}

        if o.get("domainId") and o["domainId"] in deleted_domain_ids:
            to_delete.append(tagged)
            logger.info(
                "[%s] Order %d on Deleted domain %s (%s) — will be removed",
                db_name, o["order_id"], o.get("domainId") or "?", o.get("wp_domain", ""),
            )
            continue

        if _is_replace_failure(o["status"]):
            replace_always_reset.append(tagged)
            reset_count_val = _get_reset_count_for_order(audit_session, db_name, o["order_id"])
            if reset_count_val >= REPLACE_THRESHOLD - 1:
                replace_notify.append(tagged)
                logger.warning(
                    "[%s] failed_replace order %d domain %s (%s) — reset #%d, threshold reached",
                    db_name, o["order_id"], o.get("domainId") or "?",
                    o.get("wp_domain", ""), reset_count_val + 1,
                )
        elif _is_connection_or_image_failure(o["status"]):
            if key in previously_reset:
                supplier_draft_orders.append(tagged)
                logger.warning(
                    "[%s] Repeat failure: order %d domain %s (%s) status=%s — supplier draft needed",
                    db_name, o["order_id"], o.get("domainId") or "?",
                    o.get("wp_domain", ""), o["status"],
                )
            else:
                to_reset.append(tagged)
        else:
            if key in previously_reset:
                supplier_draft_orders.append(tagged)
            else:
                to_reset.append(tagged)

    all_to_reset = to_reset + replace_always_reset

    # Delete orders on Deleted-label domains
    delete_count = 0
    if to_delete and not dry_run:
        del_ids = [o["order_id"] for o in to_delete]
        placeholders = ",".join([str(i) for i in del_ids])
        source_session.execute(text(
            f"DELETE FROM openorder WHERE id IN ({placeholders})"
        ))
        source_session.commit()
        delete_count = len(del_ids)
        logger.info("[%s] Deleted %d orders on Deleted-label domains", db_name, delete_count)
    elif to_delete:
        logger.info("[%s] [DRY-RUN] Would delete %d orders on Deleted-label domains", db_name, len(to_delete))

    # Reset remaining orders to pending
    reset_count = 0
    if all_to_reset and not dry_run:
        ids = [o["order_id"] for o in all_to_reset]
        placeholders = ",".join([str(i) for i in ids])
        source_session.execute(text(
            f"UPDATE openorder SET status = 'pending' WHERE id IN ({placeholders})"
        ))
        source_session.commit()
        reset_count = len(ids)
        logger.info("[%s] Reset %d failed orders to pending", db_name, reset_count)
    elif all_to_reset:
        logger.info("[%s] [DRY-RUN] Would reset %d orders to pending", db_name, len(all_to_reset))

    for o in to_delete:
        audit_session.add(AuditLog(
            run_id=run_id, level="INFO",
            message=f"[{db_name}] Deleted order {o['order_id']} (domain {o.get('domainId')}: {o.get('wp_domain','')}) — domain has Deleted label",
            context_json={"order_id": o["order_id"], "old_status": o["status"], "db": db_name, "action": "deleted_label79", "domainId": o.get("domainId"), "wp_domain": o.get("wp_domain", "")},
            created_at=datetime.utcnow(),
        ))

    for o in all_to_reset:
        audit_session.add(AuditLog(
            run_id=run_id, level="INFO",
            message=f"[{db_name}] Reset order {o['order_id']} (domain {o.get('domainId')}: {o.get('wp_domain','')}) from '{o['status']}' to 'pending'",
            context_json={"order_id": o["order_id"], "old_status": o["status"], "db": db_name, "dry_run": dry_run, "domainId": o.get("domainId"), "wp_domain": o.get("wp_domain", "")},
            created_at=datetime.utcnow(),
        ))

    for o in supplier_draft_orders:
        audit_session.add(AuditLog(
            run_id=run_id, level="WARNING",
            message=f"[{db_name}] REPEAT: order {o['order_id']} domain {o.get('domainId')} ({o.get('wp_domain','')}) failed again: {o['status']} — supplier draft queued",
            context_json={"order_id": o["order_id"], "status": o["status"], "db": db_name, "domainId": o.get("domainId"), "wp_domain": o.get("wp_domain", "")},
            created_at=datetime.utcnow(),
        ))

    for o in replace_notify:
        audit_session.add(AuditLog(
            run_id=run_id, level="WARNING",
            message=f"[{db_name}] REPLACE_THRESHOLD: order {o['order_id']} domain {o.get('domainId')} ({o.get('wp_domain','')}) failed_replace >= {REPLACE_THRESHOLD}x",
            context_json={"order_id": o["order_id"], "status": o["status"], "db": db_name, "domainId": o.get("domainId"), "wp_domain": o.get("wp_domain", "")},
            created_at=datetime.utcnow(),
        ))

    audit_session.commit()

    logger.info(
        "[%s] Summary: %d total failed | %d deleted (Deleted label) | %d reset to pending | "
        "%d supplier drafts | %d failed_replace (always reset) | %d failed_replace at threshold",
        db_name, len(failed_orders), delete_count if not dry_run else 0,
        reset_count if not dry_run else 0,
        len(supplier_draft_orders), len(replace_always_reset), len(replace_notify),
    )

    return {
        "db": db_name,
        "total_failed": len(failed_orders),
        "reset": reset_count if not dry_run else 0,
        "deleted": delete_count if not dry_run else 0,
        "would_reset": len(all_to_reset) if dry_run else 0,
        "supplier_draft_orders": supplier_draft_orders,
        "replace_notify_orders": replace_notify,
    }


# ---------------------------------------------------------------------------
# Supplier email draft builder
# ---------------------------------------------------------------------------

def _resolve_supplier_email(label_names: str) -> str:
    """Route to supplier email based on label names."""
    for label_key, email in SUPPLIER_ROUTING.items():
        if label_key.lower() in (label_names or "").lower():
            return email
    return DEFAULT_SUPPLIER_EMAIL


def _group_by_supplier(orders: list[dict]) -> dict[str, list[dict]]:
    """Group orders by resolved supplier email."""
    groups: dict[str, list[dict]] = {}
    for o in orders:
        supplier = _resolve_supplier_email(o.get("label_names", ""))
        groups.setdefault(supplier, []).append(o)
    return groups


def deduplicate_by_domain(orders: list[dict]) -> list[dict]:
    """Deduplicate by domainId, keeping the most recent order per domain."""
    seen: dict[str, dict] = {}
    for o in orders:
        key = f"{o.get('db', '')}:{o.get('domainId', o['order_id'])}"
        if key not in seen:
            seen[key] = {
                "domainId": o.get("domainId"),
                "wp_domain": o.get("wp_domain", ""),
                "db": o.get("db", ""),
                "label_ids": o.get("label_ids", ""),
                "label_names": o.get("label_names", ""),
                "order_count": 1,
                "statuses": {o["status"]},
                "latest_order_id": o["order_id"],
                "latest_added_on": o.get("addedOn", ""),
            }
        else:
            seen[key]["order_count"] += 1
            seen[key]["statuses"].add(o["status"])
            if str(o.get("addedOn", "")) > str(seen[key]["latest_added_on"]):
                seen[key]["latest_order_id"] = o["order_id"]
                seen[key]["latest_added_on"] = o.get("addedOn", "")

    result = []
    for v in seen.values():
        v["statuses"] = ", ".join(sorted(v["statuses"]))
        result.append(v)

    return sorted(result, key=lambda x: x.get("db", "") + str(x.get("domainId", 0)))


def build_supplier_draft_html(supplier_email: str, orders: list[dict]) -> str:
    """Build an HTML email body for the supplier about failing domains."""
    deduped = deduplicate_by_domain(orders)

    error_descriptions = {
        "failed_domain": "Website niet bereikbaar (connectie / login)",
        "failed_domain_category": "Website niet bereikbaar (connectie / categorie)",
        "failed_image": "Afbeelding uploaden geblokkeerd",
    }

    rows_html = ""
    for d in deduped:
        statuses = d.get("statuses", "")
        error_desc = " / ".join(
            error_descriptions.get(s.strip(), s.strip())
            for s in statuses.split(", ")
        )
        rows_html += (
            f"<tr>"
            f"<td>{d.get('wp_domain', '')}</td>"
            f"<td>{error_desc}</td>"
            f"<td>{d.get('order_count', 0)}</td>"
            f"</tr>\n"
        )

    recipient_name = "Hugo" if "hugo" in supplier_email else "Stefan"

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
body {{ font-family: Arial, sans-serif; font-size: 14px; color: #333; line-height: 1.6; }}
table {{ border-collapse: collapse; width: 100%; margin: 16px 0; }}
th, td {{ border: 1px solid #ddd; padding: 8px 12px; text-align: left; font-size: 13px; }}
th {{ background: #f5f5f5; font-weight: 600; }}
</style></head><body>
<p>Hoi {recipient_name},</p>

<p>Via ons geautomatiseerd systeem hebben we geconstateerd dat bij onderstaande websites herhaaldelijk
fouten optreden bij het plaatsen van linkbuilding-orders. De orders zijn meerdere keren opnieuw
geprobeerd, maar blijven falen.</p>

<p>Hieronder een overzicht van de betreffende websites en het type fout:</p>

<table>
<thead><tr><th>Website</th><th>Foutmelding</th><th>Aantal orders</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>

<p>Zou je kunnen kijken of deze problemen opgelost kunnen worden? Denk aan:</p>
<ul>
<li>Connectieproblemen of inloggegevens die niet meer werken</li>
<li>Blokkades op het uploaden van afbeeldingen</li>
</ul>

<p>Alvast bedankt voor het oppakken!</p>

<p>Met vriendelijke groet,<br>
Ruben van Melsen<br>
Blauwe Monsters</p>

<p style="color:#999;font-size:11px;">Dit bericht is automatisch gegenereerd door Domain Cleanup. Antwoord gaat naar ruben.van.melsen@blauwemonsters.nl.</p>
</body></html>"""

    return html


def build_supplier_xlsx(orders: list[dict]) -> bytes:
    """Build an xlsx matching the supplier email table."""
    deduped = deduplicate_by_domain(orders)

    error_descriptions = {
        "failed_domain": "Website niet bereikbaar (connectie / login)",
        "failed_domain_category": "Website niet bereikbaar (connectie / categorie)",
        "failed_image": "Afbeelding uploaden geblokkeerd",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Linkbuilding fouten"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    headers = ["Website", "Foutmelding", "Aantal orders"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, d in enumerate(deduped, 2):
        statuses = d.get("statuses", "")
        error_desc = " / ".join(
            error_descriptions.get(s.strip(), s.strip())
            for s in statuses.split(", ")
        )
        ws.cell(row=row_idx, column=1, value=d.get("wp_domain", ""))
        ws.cell(row=row_idx, column=2, value=error_desc)
        ws.cell(row=row_idx, column=3, value=d.get("order_count", 0))

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_supplier_drafts(all_results: list[dict]) -> list[dict]:
    """Build draft entries for each supplier from all reset results.

    Returns list of dicts ready to insert into audit_email_drafts, each
    including xlsx_bytes for attachment.
    """
    all_supplier_orders: list[dict] = []
    for r in all_results:
        all_supplier_orders.extend(r.get("supplier_draft_orders", []))

    if not all_supplier_orders:
        return []

    grouped = _group_by_supplier(all_supplier_orders)
    drafts = []

    for supplier_email, orders in grouped.items():
        deduped = deduplicate_by_domain(orders)
        html = build_supplier_draft_html(supplier_email, orders)
        xlsx_bytes = build_supplier_xlsx(orders)
        drafts.append({
            "to": supplier_email,
            "cc": RUBEN_EMAIL,
            "reply_to": RUBEN_EMAIL,
            "subject": f"Linkbuilding fouten: {len(deduped)} websites met herhaaldelijke problemen",
            "body_html": html,
            "order_count": len(orders),
            "domain_count": len(deduped),
            "orders": orders,
            "xlsx_bytes": xlsx_bytes,
        })

    return drafts


# ---------------------------------------------------------------------------
# Replace-threshold notification
# ---------------------------------------------------------------------------

def build_replace_threshold_email(all_results: list[dict]) -> str | None:
    """Build HTML for failed_replace orders that hit the threshold."""
    all_notify = []
    for r in all_results:
        all_notify.extend(r.get("replace_notify_orders", []))

    if not all_notify:
        return None

    deduped = deduplicate_by_domain(all_notify)

    rows_html = ""
    for d in deduped:
        rows_html += (
            f"<tr>"
            f"<td>{d.get('domainId', '')}</td>"
            f"<td>{d.get('wp_domain', '')}</td>"
            f"<td>{d.get('db', '')}</td>"
            f"<td>{d.get('order_count', 0)}</td>"
            f"<td>{d.get('latest_order_id', '')}</td>"
            f"</tr>\n"
        )

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
body {{ font-family: Arial, sans-serif; font-size: 14px; color: #333; }}
table {{ border-collapse: collapse; width: 100%; margin: 16px 0; }}
th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; font-size: 13px; }}
th {{ background: #f5f5f5; font-weight: 600; }}
h2 {{ color: #e67e22; }}
</style></head><body>
<h2>failed_replace — orders die {REPLACE_THRESHOLD}x of vaker zijn gefaald</h2>
<p>De volgende orders hebben herhaaldelijk een <code>failed_replace</code> status.
De ankers kunnen niet worden vervangen. Handmatige actie is vereist.</p>
<p><strong>{len(deduped)} unieke domeinen</strong> met in totaal <strong>{len(all_notify)} orders</strong>.</p>
<table>
<thead><tr><th>Domain ID</th><th>wp_domain</th><th>Database</th><th>Orders</th><th>Laatste Order</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>
<p style="color:#999;font-size:12px;">Automatisch gegenereerd door Domain Cleanup.</p>
</body></html>"""


# ---------------------------------------------------------------------------
# XLSX report
# ---------------------------------------------------------------------------

def build_xlsx_report(deduped: list[dict], output_dir: Path | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Repeat Offenders"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    headers = ["Domain ID", "wp_domain", "Database", "Label IDs", "Label Names", "Fail Statuses", "Order Count", "Latest Order ID", "Latest Date"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, d in enumerate(deduped, 2):
        ws.cell(row=row_idx, column=1, value=d.get("domainId"))
        ws.cell(row=row_idx, column=2, value=d.get("wp_domain", ""))
        ws.cell(row=row_idx, column=3, value=d.get("db", ""))
        ws.cell(row=row_idx, column=4, value=d.get("label_ids", ""))
        ws.cell(row=row_idx, column=5, value=d.get("label_names", ""))
        ws.cell(row=row_idx, column=6, value=d.get("statuses", ""))
        ws.cell(row=row_idx, column=7, value=d.get("order_count", 0))
        ws.cell(row=row_idx, column=8, value=d.get("latest_order_id"))
        ws.cell(row=row_idx, column=9, value=d.get("latest_added_on", ""))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.read()

    if output_dir:
        path = output_dir / "reports" / f"repeat_offenders_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(data)
        logger.info("Repeat offender xlsx saved to %s", path)

    return data


# ---------------------------------------------------------------------------
# Save supplier drafts + send replace notifications
# ---------------------------------------------------------------------------

def save_supplier_drafts_to_audit(
    audit_session: Session,
    supplier_drafts: list[dict],
    run_id: int | None = None,
    output_dir: Path | None = None,
) -> list[int]:
    """Store supplier drafts in audit_email_drafts + xlsx on disk. Returns list of draft IDs."""
    from app.clients.audit_db import AuditEmailDraft

    draft_ids = []
    for d in supplier_drafts:
        xlsx_path = ""
        if d.get("xlsx_bytes") and output_dir:
            xlsx_dir = output_dir / "reports"
            xlsx_dir.mkdir(parents=True, exist_ok=True)
            safe_name = d["to"].replace("@", "_at_").replace(".", "_")
            xlsx_file = xlsx_dir / f"supplier_{safe_name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            xlsx_file.write_bytes(d["xlsx_bytes"])
            xlsx_path = str(xlsx_file)
            logger.info("Supplier xlsx saved to %s", xlsx_path)

        draft = AuditEmailDraft(
            run_id=run_id or 0,
            added_by=0,
            addedby_name=d["to"],
            addedby_email=d["to"],
            subject=d["subject"],
            body_html=d["body_html"],
            body_json={
                "cc": d["cc"],
                "reply_to": d["reply_to"],
                "domain_count": d["domain_count"],
                "type": "supplier_draft",
                "xlsx_path": xlsx_path,
            },
            order_count=d["order_count"],
            created_at=datetime.utcnow(),
            send_status="draft",
        )
        audit_session.add(draft)
        audit_session.flush()
        draft_ids.append(draft.id)
        logger.info(
            "Supplier draft saved: id=%d to=%s (%d domains, %d orders, xlsx=%s)",
            draft.id, d["to"], d["domain_count"], d["order_count"], bool(xlsx_path),
        )

    audit_session.commit()
    return draft_ids


def send_replace_threshold_report(settings, all_results: list[dict]) -> None:
    """Send failed_replace threshold notification directly to Ruben."""
    import httpx

    html = build_replace_threshold_email(all_results)
    if not html:
        logger.info("No failed_replace orders at threshold — nothing to send")
        return

    if not settings.mailgun_api_key:
        logger.warning("failed_replace threshold reached but Mailgun not configured")
        return

    url = f"{settings.mailgun_api_url}/v3/{settings.mailgun_domain}/messages"
    try:
        resp = httpx.post(
            url,
            auth=("api", settings.mailgun_api_key),
            data={
                "from": settings.mailgun_from,
                "to": RUBEN_EMAIL,
                "subject": "failed_replace: orders die niet gereset kunnen worden",
                "html": html,
            },
            timeout=30,
        )
        resp.raise_for_status()
        logger.info("Replace threshold report sent to %s", RUBEN_EMAIL)
    except Exception as exc:
        logger.error("Failed to send replace threshold report: %s", exc)
